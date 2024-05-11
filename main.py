import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename, factor):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        correction = cell.value * factor
        correction_cell = sheet.cell(row, 4)
        correction_cell.value = correction

    values = Reference(sheet,
                      min_row=2,
                      max_row=sheet.max_row,
                      min_col=4,
                      max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e1')

    wb.save(filename)


process_workbook(input("File Name with [.xlsx] At The End: "), float(input("Correction Factor: ")))
